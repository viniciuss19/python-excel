import openpyxl

book = openpyxl.Workbook()

print(book.sheetnames)

book.create_sheet('Frutas')


pagfrutas = book['Frutas']
pagfrutas.append(['Maçã',"2", "R$ 3,00"])
pagfrutas.append(['Banana',"2", "R$ 1,00"])
pagfrutas.append(['Limão',"1", "R$ 1,00"])
pagfrutas.append(['Kiwi',"1", "R$ 2,00"])
pagfrutas.append(['Uva',"2", "R$ 1,50"])

book.save('Planilha de Frutas.xlsx')
